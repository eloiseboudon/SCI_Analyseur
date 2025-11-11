#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module d'export pour l'analyse financière SCI
Génère des fichiers Excel et PDF professionnels
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')


class ExporteurSCI:
    """Classe pour exporter les analyses SCI en Excel et PDF"""
    
    def __init__(self, sci):
        """
        Initialise l'exporteur avec une SCI
        
        Args:
            sci: Instance de la classe SCI
        """
        self.sci = sci
        self.workbook = None
        
    def generer_excel_complet(self, nom_fichier: str = None, duree_annees: int = 20):
        """
        Génère un fichier Excel complet avec tous les onglets d'analyse
        
        Args:
            nom_fichier: Nom du fichier de sortie (sans extension)
            duree_annees: Durée de la projection en années
        """
        if nom_fichier is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nom_fichier = f"Analyse_SCI_{self.sci.nom.replace(' ', '_')}_{timestamp}"
        
        # Créer le workbook
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Supprimer la feuille par défaut
        
        # Générer les différents onglets
        print("📝 Génération de l'analyse Excel complète...")
        
        self._creer_onglet_synthese_generale()
        self._creer_onglet_synthese_biens()
        self._creer_onglet_projection_financiere(duree_annees)
        self._creer_onglet_compte_resultat(duree_annees)
        self._creer_onglet_tresorerie(duree_annees)
        
        # Créer un onglet pour chaque bien
        for bien in self.sci.biens:
            self._creer_onglet_bien_detail(bien)
            if bien.credit:
                self._creer_onglet_credit_bien(bien)
        
        self._creer_onglet_graphiques(duree_annees)
        
        # Sauvegarder
        chemin_complet = f"/mnt/user-data/outputs/{nom_fichier}.xlsx"
        self.workbook.save(chemin_complet)
        print(f"✅ Fichier Excel créé: {nom_fichier}.xlsx")
        
        return chemin_complet
    
    def _creer_onglet_synthese_generale(self):
        """Crée l'onglet de synthèse générale"""
        ws = self.workbook.create_sheet("📊 Synthèse Générale")
        
        # Titre
        ws['A1'] = f"SYNTHÈSE GÉNÉRALE - {self.sci.nom}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Informations SCI
        ws[f'A{row}'] = "INFORMATIONS SCI"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        data_sci = [
            ["Nom de la SCI:", self.sci.nom],
            ["Année de création:", self.sci.annee_creation],
            ["Capital social:", f"{self.sci.capital_social:,.0f} €"],
            ["Nombre d'associés:", self.sci.nombre_associes],
            ["CRL (%):", f"{self.sci.crl_taux*100:.1f}%"],
            ["Frais comptable annuel:", f"{self.sci.frais_comptable_annuel:,.0f} €"],
            ["Frais bancaire annuel:", f"{self.sci.frais_bancaire_annuel:,.0f} €"],
        ]
        
        for libelle, valeur in data_sci:
            ws[f'A{row}'] = libelle
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = valeur
            row += 1
        
        row += 1
        
        # Synthèse des biens
        ws[f'A{row}'] = "SYNTHÈSE DES BIENS IMMOBILIERS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = f"Nombre de biens:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = len(self.sci.biens)
        
        row += 1
        total_prix = sum(bien.prix_total for bien in self.sci.biens)
        ws[f'A{row}'] = f"Investissement total:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{total_prix:,.0f} €"
        
        row += 1
        total_revenus = sum(bien.revenus_annuels for bien in self.sci.biens)
        ws[f'A{row}'] = f"Revenus locatifs annuels:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{total_revenus:,.0f} €"
        
        row += 1
        total_charges = sum(bien.charges_annuelles for bien in self.sci.biens)
        ws[f'A{row}'] = f"Charges annuelles (biens):"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{total_charges:,.0f} €"
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _creer_onglet_synthese_biens(self):
        """Crée l'onglet de synthèse des biens"""
        ws = self.workbook.create_sheet("🏢 Biens Immobiliers")
        
        # Générer le DataFrame de synthèse
        synthese = self.sci.generer_synthese_biens()
        
        # Écrire le DataFrame dans la feuille
        for r_idx, row in enumerate(dataframe_to_rows(synthese, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style pour l'en-tête
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _creer_onglet_projection_financiere(self, duree_annees: int):
        """Crée l'onglet de projection financière"""
        ws = self.workbook.create_sheet("📈 Projection Financière")
        
        # Générer la projection
        projection = self.sci.generer_projection(duree_annees)
        
        # Sélectionner les colonnes à afficher
        cols = [
            'annee', 'revenus_locatifs', 'charges_exploitation',
            'amortissements', 'interets_credits', 'resultat_avant_impot',
            'impot_societes', 'resultat_net', 'cashflow', 'reserves_fin'
        ]
        
        projection_export = projection[cols].copy()
        projection_export.columns = [
            'Année', 'Revenus Locatifs', 'Charges Exploitation',
            'Amortissements', 'Intérêts Crédits', 'Résultat Av. IS',
            'Impôt Sociétés', 'Résultat Net', 'Cash-Flow', 'Réserves'
        ]
        
        # Écrire dans la feuille
        for r_idx, row in enumerate(dataframe_to_rows(projection_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Format numérique pour les colonnes de valeurs
                    if c_idx > 1:
                        cell.number_format = '#,##0 €'
        
        # Ajuster les largeurs
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
    
    def _creer_onglet_compte_resultat(self, duree_annees: int):
        """Crée l'onglet du compte de résultat"""
        ws = self.workbook.create_sheet("💰 Compte de Résultat")
        
        # Générer la projection
        projection = self.sci.generer_projection(duree_annees)
        
        cols = [
            'annee', 'revenus_locatifs', 'charges_exploitation',
            'frais_exceptionnels', 'amortissements', 'resultat_exploitation',
            'interets_credits', 'resultat_avant_impot', 'impot_societes', 'resultat_net'
        ]
        
        cr = projection[cols].copy()
        cr.columns = [
            'Année', 'Revenus Locatifs', 'Charges Exploitation',
            'Frais Exceptionnels', 'Amortissements', 'Résultat Exploitation',
            'Intérêts Crédits', 'Résultat Av. IS', 'IS', 'Résultat Net'
        ]
        
        # Écrire dans la feuille
        for r_idx, row in enumerate(dataframe_to_rows(cr, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx > 1:
                        cell.number_format = '#,##0 €'
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
    
    def _creer_onglet_tresorerie(self, duree_annees: int):
        """Crée l'onglet de trésorerie"""
        ws = self.workbook.create_sheet("💵 Trésorerie")
        
        projection = self.sci.generer_projection(duree_annees)
        
        cols = [
            'annee', 'encaissements', 'decaissements', 'mensualites_credit',
            'cashflow', 'tresorerie_realisee', 'resultat_net', 'reserves_fin'
        ]
        
        tresorerie = projection[cols].copy()
        tresorerie.columns = [
            'Année', 'Encaissements', 'Décaissements', 'Mensualités Crédit',
            'Cash-Flow', 'Trésorerie Réalisée', 'Résultat Net', 'Réserves'
        ]
        
        for r_idx, row in enumerate(dataframe_to_rows(tresorerie, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx > 1:
                        cell.number_format = '#,##0 €'
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
    
    def _creer_onglet_bien_detail(self, bien):
        """Crée un onglet détaillé pour un bien"""
        ws = self.workbook.create_sheet(f"🏠 {bien.nom[:20]}")
        
        # Titre
        ws['A1'] = f"DÉTAIL - {bien.nom}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Informations générales
        data = [
            ["INFORMATIONS GÉNÉRALES", ""],
            ["Année d'achat:", bien.annee_achat],
            ["Prix d'achat:", f"{bien.prix_achat:,.0f} €"],
            ["Frais d'agence:", f"{bien.frais_agence:,.0f} €"],
            ["Frais de notaire:", f"{bien.frais_notaire:,.0f} €"],
            ["Travaux:", f"{bien.travaux:,.0f} €"],
            ["Meubles:", f"{bien.meubles:,.0f} €"],
            ["Prix total:", f"{bien.prix_total:,.0f} €"],
            ["", ""],
            ["FINANCEMENT", ""],
            ["Apport SCI:", f"{bien.apport_sci:,.0f} €"],
        ]
        
        if bien.credit:
            data.extend([
                ["Capital emprunté:", f"{bien.credit.capital_emprunte:,.0f} €"],
                ["Taux annuel:", f"{bien.credit.taux_annuel*100:.2f}%"],
                ["Durée:", f"{bien.credit.duree_annees} ans"],
                ["Mensualité:", f"{bien.credit.calculer_mensualite():,.2f} €"],
                ["Total intérêts:", f"{bien.credit.calculer_total_interets():,.0f} €"],
            ])
        
        data.extend([
            ["", ""],
            ["REVENUS LOCATIFS", ""],
            ["Nombre d'appartements:", len(bien.appartements)],
            ["Revenus mensuels:", f"{bien.revenus_mensuels:,.2f} €"],
            ["Revenus annuels:", f"{bien.revenus_annuels:,.0f} €"],
            ["", ""],
            ["CHARGES ANNUELLES", ""],
            ["Taxe foncière:", f"{bien.taxe_fonciere:,.0f} €"],
            ["Charges copropriété:", f"{bien.charges_copro:,.0f} €"],
            ["Autres charges:", f"{bien.autres_charges:,.2f} €"],
            ["Total charges:", f"{bien.charges_annuelles:,.2f} €"],
            ["", ""],
            ["RENTABILITÉ", ""],
            ["Rentabilité brute:", f"{bien.calculer_rentabilite_brute():.2f}%"],
            ["Rentabilité nette:", f"{bien.calculer_rentabilite_nette():.2f}%"],
        ])
        
        for libelle, valeur in data:
            ws[f'A{row}'] = libelle
            if libelle and not valeur:  # C'est un titre de section
                ws[f'A{row}'].font = Font(bold=True, size=11)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                ws.merge_cells(f'A{row}:D{row}')
            else:
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = valeur
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _creer_onglet_credit_bien(self, bien):
        """Crée un onglet pour le tableau d'amortissement du crédit"""
        ws = self.workbook.create_sheet(f"💳 Crédit {bien.nom[:15]}")
        
        # Générer le tableau d'amortissement
        tableau = bien.credit.generer_tableau_amortissement()
        
        if tableau.empty:
            return
        
        # Ajouter une colonne Année
        tableau['Année'] = ((tableau['Mois'] - 1) // 12) + 1
        
        # Réorganiser les colonnes
        cols_ordre = ['Mois', 'Année', 'Capital restant début', 'Mensualité', 
                     'Intérêts', 'Capital amorti', 'Capital restant fin']
        tableau = tableau[cols_ordre]
        
        # Écrire dans la feuille
        for r_idx, row in enumerate(dataframe_to_rows(tableau, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx > 2:  # Colonnes numériques
                        cell.number_format = '#,##0.00 €'
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
    
    def _creer_onglet_graphiques(self, duree_annees: int):
        """Crée un onglet avec un résumé graphique"""
        ws = self.workbook.create_sheet("📊 Graphiques")
        
        # Note : La création de graphiques avec openpyxl est complexe
        # Pour l'instant, on crée juste un tableau récapitulatif
        
        ws['A1'] = "RÉSUMÉ GRAPHIQUE"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Les graphiques détaillés peuvent être générés dans Excel en utilisant les données des autres onglets."
        ws.merge_cells('A3:D3')


if __name__ == "__main__":
    # Ce module est conçu pour être importé
    # Pour un exemple d'utilisation, voir generate_report.py
    pass
